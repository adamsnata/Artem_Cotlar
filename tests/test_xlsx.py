import os.path
from openpyxl import load_workbook

# TODO оформить в тест, добавить ассерты и использовать универсальный путь
current = os.path.abspath(__file__)
print(current)
project_root = os.path.dirname(current)
print(project_root)
joined_path = os.path.join(project_root, 'resources', 'file_example_XLSX_50.xlsx')
print(joined_path)



def test_xlsx():
    workbook = load_workbook(joined_path)
    sheet = workbook.active
    cel_rows = sheet.cell(row=3, column=2).value
    print(f"Пересечение строк: {cel_rows}")

    assert cel_rows == 'Mara'
